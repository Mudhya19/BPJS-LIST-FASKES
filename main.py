import openpyxl
from openpyxl.utils import get_column_letter

# Buka workbook dan pilih worksheet
file_path = 'data\juli 2024.xlsx'  # Ganti dengan nama file Excel Anda
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Fungsi untuk menambahkan kolom kosong di antara kolom
def add_spacing_columns(ws):
    # Dapatkan jumlah kolom terakhir dengan data
    last_col = ws.max_column
    
    # Sisipkan kolom kosong setelah setiap kolom
    for col in range(last_col, 1, -1):
        ws.insert_cols(col)

# Fungsi untuk mengatur lebar kolom otomatis (AutoFit)
def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Dapatkan nomor kolom (1, 2, 3, ...)
        column_letter = get_column_letter(column)  # Ubah ke format huruf (A, B, C, ...)
        for cell in col:
            try:
                # Ukur panjang konten sel
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Tambahkan sedikit ruang
        ws.column_dimensions[column_letter].width = adjusted_width

# Tambahkan kolom kosong di antara kolom yang ada
add_spacing_columns(ws)

# Atur lebar kolom agar sesuai dengan data
autofit_columns(ws)

# Simpan perubahan ke file Excel baru
wb.save('data_dengan_jarak_dan_autofit.xlsx')

print("Proses selesai. File telah disimpan sebagai 'data_dengan_jarak_dan_autofit.xlsx'")
